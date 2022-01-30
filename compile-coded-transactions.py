import openpyxl
import os
import time

target_dir = 'C:\\Users\\liamh\\OneDrive\\Documents\\Banking\\19-20\\'
year_workbook = target_dir + 'Year Coded Summary 19-20.xlsx'

def main():
    # get files in directory
    files = os.listdir(target_dir)

    copy_coded_transactions(files)

    return

def copy_coded_transactions(files):

    # for each file in directory
    for file in files:
        year_summary_book = openpyxl.load_workbook(year_workbook)
        if file != "Year Coded Summary 19-20.xlsx":
            # open month spreadsheet
            try:
                month_book = openpyxl.load_workbook(target_dir + file)
            except openpyxl.utils.exceptions.InvalidFileException:
                pass
            sheet_names = month_book.sheetnames
            # get raw transactions sheet
            raw_trans_sheet_month = month_book.get_sheet_by_name(sheet_names[1])
            # get raw max row and column
            mr = raw_trans_sheet_month.max_row
            mc = raw_trans_sheet_month.max_column

            # get max raw year row
            year_sheet_names = year_summary_book.sheetnames
            raw_year_sheet = year_summary_book.get_sheet_by_name(year_sheet_names[0])
            mr_year = raw_year_sheet.max_row
            # create pointer to track which row of the year sheet to populate
            # we want to append new data to existing data from previous months
            year_row_pointer = mr_year + 1
            print('year_row_pointer_start_value:' + str(year_row_pointer))


            # loop through coded transactions
            # starting from second column to avoid headers
            for i in range(2, mr + 1):

                for j in range(1, 5):
                    month_cell = raw_trans_sheet_month.cell(row = i, column = j)
                    raw_year_sheet.cell(row = year_row_pointer, column = j).value = month_cell.value
                # increment year_row_pointer
                year_row_pointer = year_row_pointer + 1


            month_book.close()
            year_summary_book.save(year_workbook)
            year_summary_book.close()
            print('year_row_pointer_end_value:' + str(year_row_pointer))
    return

main()