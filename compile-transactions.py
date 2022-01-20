
import openpyxl
import os

target_dir = 'C:\\Users\\liamh\\OneDrive\\Documents\\Banking\\19-20\\'

# get files in directory
files = os.listdir(target_dir)

# open summary file
summary_book = openpyxl.load_workbook(target_dir + 'Year Summary 19-20.xlsx')
# get raw sheet
# raw_trans_sheet = summary_book.get_sheet_by_name('Raw')
# get coded sheet
# coded_sheet = summary_book.get_sheet_by_name('Coded')


# for each file in directory
for file in files:
    if file != "Year Summary 19-20.xlsx":
        # open month spreadsheet
        month_book = openpyxl.load_workbook(target_dir + file)
        sheet_names = month_book.sheetnames
        # copy raw transactions
        raw_trans_sheet_month = month_book.get_sheet_by_name(sheet_names[0])
        # get max row and column
        mr = raw_trans_sheet_month.max_row
        mc = raw_trans_sheet_month.max_column
        # loop through raw transactions 
        # close wb
        month_book.close()

        print(mr)

summary_book.close()


